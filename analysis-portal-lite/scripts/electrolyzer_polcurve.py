#!/usr/bin/env python3
"""
PEM Electrolyzer Polarization Curve Analyzer
=============================================
Loads raw potentiostat CSV data (e.g. Biologic, Gamry, Squidstat),
auto-detects voltage and current columns, extracts representative
data from each dwell period, identifies polcurve cycles, and plots
each cycle as a separate V vs j curve.

Usage:
  python electrolyzer_polcurve.py                  # interactive
  python electrolyzer_polcurve.py --file data.csv  # direct
"""
def run(input_dir: str, output_dir: str) -> dict:
    # Read CSVs from input_dir
    # Write PNGs and Excel files to output_dir
    # Return a summary dict
import subprocess, sys


def _ensure_deps():
    for pkg in ['numpy', 'matplotlib', 'scipy', 'openpyxl']:
        try: __import__(pkg)
        except ImportError:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg],
                                  stdout=subprocess.DEVNULL)

_ensure_deps()

import argparse, csv, gc
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from pathlib import Path


# ═══════════════════════════════════════════════════════════════════
#  Column detection
# ═══════════════════════════════════════════════════════════════════

def _clean_path(p):
    p = p.strip()
    if p.startswith('& '): p = p[2:]
    return p.strip().strip('"').strip("'").strip('\u2018\u2019\u201c\u201d\u202a\u200b')


def _match_col(header, candidates):
    """Return the first column name in header that matches any candidate (case-insensitive substring)."""
    hl = header.lower()
    for c in candidates:
        if c in hl:
            return True
    return False


def detect_columns(fieldnames):
    """
    Auto-detect voltage, current, step, and repeat columns.

    Returns dict with keys: 'v_col', 'i_col', 'step_col', 'repeat_col', 'step_name_col'
    (values are column name strings or None).
    """
    result = {k: None for k in ('v_col', 'i_col', 'step_col', 'repeat_col', 'step_name_col', 'time_col')}

    v_candidates = ['working electrode (v)', 'voltage (v)', 'v_cell', 'vcell',
                     'e_stack', 'ewe (v)', 'potential (v)', 'working electrode vs']
    # Match "Current (A)" but not "Current Density"
    i_candidates = ['current (a)', 'i (a)', 'current(a)']
    step_candidates = ['step number', 'step_number', 'step no', 'ns']
    repeat_candidates = ['repeats', 'repeat', 'cycle number', 'cycle']
    step_name_candidates = ['step name', 'step_name', 'technique']
    time_candidates = ['elapsed time', 'time (s)', 'elapsed_time', 'time(s)']

    for fn in fieldnames:
        fl = fn.lower().strip()

        # Voltage — but skip "current density" columns that contain "a/m"
        if result['v_col'] is None:
            for c in v_candidates:
                if c in fl:
                    result['v_col'] = fn
                    break

        # Current — must NOT be "current density"
        if result['i_col'] is None and 'density' not in fl:
            for c in i_candidates:
                if c in fl:
                    result['i_col'] = fn
                    break

        if result['step_col'] is None:
            for c in step_candidates:
                if c in fl:
                    result['step_col'] = fn
                    break

        if result['repeat_col'] is None:
            for c in repeat_candidates:
                if c in fl:
                    result['repeat_col'] = fn
                    break

        if result['step_name_col'] is None:
            for c in step_name_candidates:
                if c in fl:
                    result['step_name_col'] = fn
                    break

        if result['time_col'] is None:
            for c in time_candidates:
                if c in fl:
                    result['time_col'] = fn
                    break

    return result


# ═══════════════════════════════════════════════════════════════════
#  Data loading
# ═══════════════════════════════════════════════════════════════════

def load_data(filepath):
    """
    Load CSV/TSV, keeping only the columns needed for analysis.

    Streams line-by-line to avoid loading the entire file into memory.

    Returns
    -------
    data : dict of column_name → numpy array
    fieldnames : list of all column name strings (for detection)
    """
    path = Path(_clean_path(filepath))
    if not path.exists():
        raise FileNotFoundError(
            f"File not found: {path}\n"
            f"  On Windows: right-click → Copy as path, then paste.")

    # Read header to detect delimiter and columns
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        first_line = f.readline()

    delim = '\t' if '\t' in first_line else ','
    fn = [c.strip() for c in first_line.strip().split(delim)]

    # Detect which columns we actually need
    cols = detect_columns(fn)
    keep = set(v for v in cols.values() if v is not None)

    # Map needed column names to their index in each row
    col_idx = {name: i for i, name in enumerate(fn) if name in keep}

    # Stream file, only storing needed columns
    raw = {name: [] for name in col_idx}
    n_rows = 0

    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        next(f)  # skip header
        for line in f:
            parts = line.strip().split(delim)
            if len(parts) < max(col_idx.values()) + 1:
                continue
            for name, idx in col_idx.items():
                raw[name].append(parts[idx].strip())
            n_rows += 1

    # Convert to numpy
    data = {}
    for name, vals in raw.items():
        try:
            data[name] = np.array([float(x) if x.lower() != 'nan' else np.nan
                                   for x in vals])
        except ValueError:
            data[name] = np.array(vals)
    del raw

    print(f"  Loaded {n_rows:,} rows from '{path.name}'  ({len(col_idx)} of {len(fn)} columns kept)")

    return data, fn


# ═══════════════════════════════════════════════════════════════════
#  Dwell extraction
# ═══════════════════════════════════════════════════════════════════

def extract_dwells_from_steps(data, cols, geo_area):
    """
    Use step/repeat columns to define dwells. Extract representative
    (V, j) from the stable tail of each dwell.

    Returns list of dicts: [{'V': float, 'j': float, 'step': int, 'repeat': int}, ...]
    """
    V_raw = data[cols['v_col']]
    I_raw = data[cols['i_col']]
    step = data[cols['step_col']].astype(int)
    repeat = data[cols['repeat_col']].astype(int)
    has_time = cols['time_col'] is not None
    if has_time:
        T_raw = data[cols['time_col']]

    # Identify unique (step, repeat) segments in time order
    segments = []
    cur_key = (step[0], repeat[0])
    seg_start = 0

    for i in range(1, len(step)):
        key = (step[i], repeat[i])
        if key != cur_key:
            segments.append((cur_key, seg_start, i))
            cur_key = key
            seg_start = i
    segments.append((cur_key, seg_start, len(step)))

    # Determine which step names to include for polcurve
    polcurve_step_names = None
    if cols['step_name_col'] is not None:
        names = data[cols['step_name_col']]
        # Collect unique step names
        unique_names = set(names)
        has_potential = any('constant potential' in n.lower() for n in unique_names)
        has_current = any('constant current' in n.lower() for n in unique_names)

        if has_potential:
            # Prefer potentiostatic setpoints for polcurve
            polcurve_step_names = {'constant potential'}
        elif has_current:
            polcurve_step_names = {'constant current'}
        # If neither, include all steps

    dwells = []
    for (s, r), start, end in segments:
        n_pts = end - start

        # Skip very short segments (< 10 pts)
        if n_pts < 10:
            continue

        # Filter by step name
        if polcurve_step_names is not None:
            sname = names[start].lower().strip()
            if not any(psn in sname for psn in polcurve_step_names):
                continue

        # Extract representative V and I from stable tail
        V_seg = V_raw[start:end]
        I_seg = I_raw[start:end]

        # Voltage: controlled variable → mean of segment
        V_sp = np.nanmean(V_seg)

        # Current: response variable → find stable tail
        # Walk backwards to find where current stabilizes
        n = len(I_seg)
        stable_start = n - 1
        for k in range(n - 2, max(0, n // 2) - 1, -1):
            tail = I_seg[k:]
            tail_clean = tail[~np.isnan(tail)]
            if len(tail_clean) < 2:
                continue
            t_mean = np.mean(tail_clean)
            t_std = np.std(tail_clean)
            if t_std > max(0.03 * abs(t_mean), 0.001):
                break
            stable_start = k

        # Average last 20 stable points
        n_tail = min(20, n - stable_start)
        tail_sl = slice(n - n_tail, n)
        I_tail = I_seg[tail_sl]
        I_tail = I_tail[~np.isnan(I_tail)]
        if len(I_tail) == 0:
            continue

        I_rep = np.mean(I_tail)
        j_rep = I_rep / geo_area

        dwells.append({
            'V': V_sp,
            'j': j_rep,
            'step': s,
            'repeat': r,
            'n_pts': n_pts,
            't_mid': float(np.nanmean(T_raw[start:end])) if has_time else None,
        })

    return dwells


def extract_dwells_generic(data, cols, geo_area):
    """
    Fallback when no step/repeat columns exist.
    Uses voltage-stability grouping (same approach as the fuel cell script).
    """
    V_raw = data[cols['v_col']]
    I_raw = data[cols['i_col']]

    # Detect voltage step size
    step_size = _detect_step(V_raw)
    if step_size is None:
        step_size = _detect_step(I_raw)

    if step_size and step_size > 0:
        af, rf = step_size * 0.25, 0.01
    else:
        ds = np.abs(np.diff(V_raw[~np.isnan(V_raw)]))
        ne = np.percentile(ds, 50) if len(ds) > 0 else 0.01
        af, rf = max(ne * 10, 0.003), 0.03

    signal = V_raw
    grps, gs, gm, gc = [], 0, signal[0], 1
    for i in range(1, len(signal)):
        if np.isnan(signal[i]):
            continue
        if gc == 0:
            gs, gm, gc = i, signal[i], 1
            continue
        mu = gm / gc
        if abs(signal[i] - mu) <= max(af, rf * abs(mu)):
            gm += signal[i]; gc += 1
        else:
            grps.append((gs, i))
            gs, gm, gc = i, signal[i], 1
    if gc > 0:
        grps.append((gs, len(signal)))

    dwells = []
    for start, end in grps:
        n = end - start
        if n < 10:
            continue

        V_sp = np.nanmean(V_raw[start:end])
        I_seg = I_raw[start:end]

        n_tail = min(20, n)
        I_tail = I_seg[n - n_tail:n]
        I_tail = I_tail[~np.isnan(I_tail)]
        if len(I_tail) == 0:
            continue

        dwells.append({
            'V': V_sp,
            'j': np.mean(I_tail) / geo_area,
            'step': 0,
            'repeat': len(dwells),
            'n_pts': n,
        })

    return dwells


def _detect_step(signal):
    """Auto-detect step size via histogram peak spacing."""
    s = signal[~np.isnan(signal)]
    if len(s) < 100:
        return None
    span = s.max() - s.min()
    if span < 1e-6:
        return None
    bw = max(span / 200, 1e-4)
    bins = np.arange(s.min() - bw, s.max() + 2 * bw, bw)
    if len(bins) < 10:
        return None
    hist, edges = np.histogram(s, bins=bins)
    ctrs = (edges[:-1] + edges[1:]) / 2
    thr = np.percentile(hist[hist > 0], 50)
    pks = [ctrs[i] for i in range(1, len(hist) - 1)
           if hist[i] > thr and hist[i] >= hist[i - 1] and hist[i] >= hist[i + 1]]
    if len(pks) < 4:
        return None
    pks = np.sort(pks)
    gaps = np.diff(pks)
    gr = np.round(gaps / bw) * bw
    ug, uc = np.unique(gr, return_counts=True)
    m = ug > bw * 0.5
    return float(ug[m][np.argmax(uc[m])]) if m.any() else None


# ═══════════════════════════════════════════════════════════════════
#  Cycle detection
# ═══════════════════════════════════════════════════════════════════

def detect_cycles(dwells):
    """
    Group dwells into polcurve cycles.

    A cycle is a sequence of dwells with monotonically changing voltage
    setpoints. A new cycle starts when:
      - Voltage drops significantly (return to low V after high V)
      - A gap in repeat numbering (indicates a separator step)

    Dwells with V below the polcurve range (e.g. recovery holds at
    ~1.25 V while the sweep goes 1.40–1.80 V) are treated as cycle
    boundaries.

    Returns list of cycles, each a list of dwell dicts sorted by V.
    """
    if not dwells:
        return []

    # Find boundary voltage by looking for the largest gap in sorted setpoints
    # This separates recovery/baseline dwells (e.g. 1.25 V) from polcurve
    # setpoints (e.g. 1.40–1.80 V)
    voltages = np.array(sorted(set(round(d['V'], 3) for d in dwells)))

    if len(voltages) < 3:
        return [sorted(dwells, key=lambda d: d['V'])]

    gaps = np.diff(voltages)
    median_gap = np.median(gaps)
    largest_gap_idx = np.argmax(gaps)
    largest_gap = gaps[largest_gap_idx]

    # Boundary gap must be significantly larger than typical step spacing
    # (at least 3× median gap AND > 50 mV absolute)
    if largest_gap > max(3.0 * median_gap, 0.05):
        v_boundary = voltages[largest_gap_idx] + largest_gap * 0.5
    else:
        v_boundary = voltages.min() - 0.01  # no boundary, keep everything

    # Determine polcurve voltage range (above boundary)
    pc_voltages = voltages[voltages > v_boundary]
    if len(pc_voltages) < 2:
        pc_voltages = voltages
    v_lo = pc_voltages.min()
    v_hi = pc_voltages.max()
    v_span = v_hi - v_lo

    # Walk through dwells, grouping into cycles.
    # A new cycle starts when:
    #   1. Voltage drops back toward the bottom of the sweep after having
    #      been in the upper half (voltage reset)
    #   2. Step number changes
    #   3. Dwell is below the boundary voltage (baseline/recovery)
    cycles = []
    current_cycle = []

    for i, d in enumerate(dwells):
        # Skip baseline/recovery dwells
        if d['V'] < v_boundary:
            if len(current_cycle) >= 3:
                cycles.append(current_cycle)
            current_cycle = []
            continue

        if current_cycle:
            prev = current_cycle[-1]

            # Different step → new cycle
            if d['step'] != prev['step']:
                if len(current_cycle) >= 3:
                    cycles.append(current_cycle)
                current_cycle = []

            # Voltage reset: V drops by more than 30% of the polcurve span
            # (e.g. from 1.80 V back to 1.40 V when span is 0.40 V)
            elif d['V'] < prev['V'] - 0.30 * v_span:
                if len(current_cycle) >= 3:
                    cycles.append(current_cycle)
                current_cycle = []

        current_cycle.append(d)

    # Flush last cycle
    if len(current_cycle) >= 3:
        cycles.append(current_cycle)

    # Sort each cycle by voltage ascending
    for cyc in cycles:
        cyc.sort(key=lambda d: d['V'])

    return cycles


# ═══════════════════════════════════════════════════════════════════
#  Plotting
# ═══════════════════════════════════════════════════════════════════

def plot_cycles(cycles, geo_area, title=None, save_path=None):
    """
    Plot all polcurve cycles on one figure.
    V on y-axis, j on x-axis, one curve per cycle.
    """
    n_cyc = len(cycles)
    if n_cyc == 0:
        print("  No cycles to plot.")
        return

    fig, ax = plt.subplots(figsize=(10, 6), dpi=120)
    cmap = plt.cm.viridis(np.linspace(0, 0.9, n_cyc))

    for i, cyc in enumerate(cycles):
        j_arr = np.array([d['j'] for d in cyc])
        V_arr = np.array([d['V'] for d in cyc])
        label = f'Cycle {i + 1}'
        ax.plot(j_arr, V_arr, 'o-', ms=4, lw=1.2, color=cmap[i], label=label)

    ax.set_xlabel('Current density  j  [A/cm²]', fontsize=12)
    ax.set_ylabel('Cell voltage  V  [V]', fontsize=12)
    ax.set_xlim(left=0)
    ax.grid(True, alpha=0.3)

    # Smart legend: if many cycles, put colorbar instead
    if n_cyc <= 15:
        ax.legend(fontsize=8, loc='upper left', ncol=max(1, n_cyc // 8))
    else:
        sm = plt.cm.ScalarMappable(cmap=plt.cm.viridis,
                                    norm=plt.Normalize(1, n_cyc))
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax, pad=0.02)
        cbar.set_label('Cycle number', fontsize=11)

    ttl = title or f'Electrolyzer Polarization  ({n_cyc} cycles, {geo_area:.1f} cm²)'
    ax.set_title(ttl, fontsize=12, fontweight='bold')

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()

    return fig


def extract_j_at_voltage(cycles, v_target, v_tol=0.015):
    """
    Extract current density at a target voltage from each cycle.

    Uses the dwell closest to v_target (within v_tol). If no dwell
    is within tolerance, that cycle returns NaN.

    Returns
    -------
    cycle_nums : array of cycle numbers (1-indexed)
    j_values   : array of j at v_target for each cycle
    """
    cycle_nums = []
    j_values = []

    for i, cyc in enumerate(cycles):
        best_j = np.nan
        best_dv = v_tol + 1  # start outside tolerance

        for d in cyc:
            dv = abs(d['V'] - v_target)
            if dv < best_dv:
                best_dv = dv
                best_j = d['j']

        if best_dv <= v_tol:
            cycle_nums.append(i + 1)
            j_values.append(best_j)

    return np.array(cycle_nums), np.array(j_values)


def detect_stabilization(cycle_nums, j_values, n_stable=5,
                         n_lookback=10, threshold_pct=1.0):
    """
    Detect the cycle at which performance stabilizes.

    At each candidate cycle i, the reference is the mean of the
    preceding `n_lookback` cycles. Stabilization is the first cycle
    where j for that cycle and the next `n_stable`−1 cycles are all
    within `threshold_pct`% of that rolling reference.

    Example with defaults: stable at cycle 20 means cycles 20–24
    are each within 1% of mean(cycles 10–19).

    Parameters
    ----------
    cycle_nums : array of cycle numbers
    j_values   : array of j at each cycle
    n_stable   : consecutive cycles that must be within band (default 3)
    n_lookback : number of preceding cycles for rolling reference (default 10)
    threshold_pct : max deviation from reference as % (default 0.5)

    Returns
    -------
    stable_cycle : int or None
    """
    n = len(j_values)
    if n < n_lookback + n_stable:
        return None

    frac = threshold_pct / 100.0

    for i in range(n_lookback, n - n_stable + 1):
        j_ref = np.mean(j_values[i - n_lookback:i])
        if abs(j_ref) < 1e-9:
            continue
        band = frac * abs(j_ref)

        # Check if cycles i, i+1, ..., i+n_stable-1 are all within band
        window = j_values[i:i + n_stable]
        if np.all(np.abs(window - j_ref) <= band):
            return int(cycle_nums[i])

    return None


def plot_j_vs_cycle(cycles, v_targets, save_path=None):
    """
    Plot current density at specified voltages vs cycle number,
    with a vertical indicator where performance stabilizes.

    Parameters
    ----------
    cycles : list of cycle dicts
    v_targets : list of voltages to track (e.g. [1.8, 1.7])
    save_path : str or None
    """
    fig, ax = plt.subplots(figsize=(9, 5), dpi=120)
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']
    markers = ['o', 's', '^', 'D', 'v']

    stable_cycles = []

    for k, vt in enumerate(v_targets):
        cn, jv = extract_j_at_voltage(cycles, vt)
        if len(cn) == 0:
            print(f"  No data at V = {vt:.3f} V")
            continue
        c = colors[k % len(colors)]
        m = markers[k % len(markers)]
        ax.plot(cn, jv, f'{m}-', color=c, ms=5, lw=1.2,
                label=f'j @ {vt:.2f} V')

        # Detect stabilization
        sc = detect_stabilization(cn, jv)
        if sc is not None:
            stable_cycles.append((vt, sc))
            # Mark the stabilization point on the curve
            sc_idx = np.where(cn == sc)[0]
            if len(sc_idx) > 0:
                ax.plot(sc, jv[sc_idx[0]], '*', color=c, ms=14,
                        markeredgecolor='k', markeredgewidth=0.8, zorder=6)
            print(f"  Stabilization at {vt:.2f} V: cycle {sc}")

    ax.set_xlabel('Cycle number', fontsize=12)
    ax.set_ylabel('Current density  j  [A/cm²]', fontsize=12)
    ax.set_xlim(left=0)
    ax.grid(True, alpha=0.3)
    ax.set_title('Current Density vs. Cycle', fontsize=12, fontweight='bold')

    handles, labels = ax.get_legend_handles_labels()
    fig.legend(handles, labels,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(handles), fontsize=9,
               frameon=True, fancybox=True)

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()

    return fig


# ═══════════════════════════════════════════════════════════════════
#  Data export
# ═══════════════════════════════════════════════════════════════════

def export_excel(filepath, cycles, v_targets=[1.8, 1.7],
                 eis_mapped=None, loss_data=None, fit_result=None,
                 eis_results=None, ir_data=None, geo_area=5.0):
    """
    Export all analysis data to a multi-sheet Excel workbook.

    Sheets:
      - Polcurve Data: cycle, V, j for every setpoint in every cycle
      - j vs Cycle: cycle number, j at each target voltage, stability info
      - HFR vs Cycle: cycle number, ASR from EIS (if available)
      - Losses vs Cycle: per-cycle fitted loss breakdown at target voltage
      - Model Fit: last cycle fit — data, model, residuals, components
      - EIS: frequency, Z', -Z'' for each EIS measurement
    """
    from openpyxl import Workbook
    wb = Workbook()

    # ── Sheet 1: Polcurve Data ──
    ws = wb.active
    ws.title = "Polcurve Data"
    ws.append(['Cycle', 'V_setpoint [V]', 'j [A/cm²]', 'Step', 'Repeat'])
    for i, cyc in enumerate(cycles):
        for d in cyc:
            ws.append([i + 1, round(d['V'], 4), round(d['j'], 6),
                       d['step'], d['repeat']])

    # ── Sheet 2: j vs Cycle ──
    ws2 = wb.create_sheet("j vs Cycle")
    header = ['Cycle']
    for vt in v_targets:
        header.append(f'j @ {vt:.2f} V [A/cm²]')
    ws2.append(header)

    # Extract j at each target voltage
    j_at_v = {}
    for vt in v_targets:
        cn, jv = extract_j_at_voltage(cycles, vt)
        j_at_v[vt] = dict(zip(cn.astype(int), jv))

    all_cycle_nums = sorted(set().union(*(j_at_v[vt].keys() for vt in v_targets)))
    for cn in all_cycle_nums:
        row = [cn]
        for vt in v_targets:
            row.append(round(j_at_v[vt].get(cn, float('nan')), 6))
        ws2.append(row)

    # Add stability info
    ws2.append([])
    ws2.append(['Stability Detection'])
    for vt in v_targets:
        cn_arr, jv_arr = extract_j_at_voltage(cycles, vt)
        sc = detect_stabilization(cn_arr, jv_arr)
        ws2.append([f'{vt:.2f} V', f'Stable @ cycle {sc}' if sc else 'Not stabilized'])

    # ── Sheet 3: HFR vs Cycle ──
    if eis_mapped:
        ws3 = wb.create_sheet("HFR vs Cycle")
        ws3.append(['Cycle', 'ASR [mΩ·cm²]', 'Elapsed Time [s]'])
        for em in eis_mapped:
            ws3.append([em['cycle'], round(em['asr_mohm_cm2'], 2),
                        round(em['t_eis'], 1) if em['t_eis'] else ''])

    # ── Sheet 4: Losses vs Cycle ──
    if loss_data is not None:
        cn_loss, j_loss, losses = loss_data
        if len(cn_loss) > 0:
            ws4 = wb.create_sheet("Losses vs Cycle")
            ws4.append(['Cycle', 'j [A/cm²]',
                        'η_anode [mV]', 'η_cathode [mV]', 'η_kinetic_total [mV]',
                        'V_ohmic [mV]', 'V_mt [mV]'])
            for i in range(len(cn_loss)):
                eta_total = losses['eta_anode_mV'][i] + losses['eta_cathode_mV'][i]
                ws4.append([
                    int(cn_loss[i]),
                    round(j_loss[i], 6),
                    round(losses['eta_anode_mV'][i], 2),
                    round(losses['eta_cathode_mV'][i], 2),
                    round(eta_total, 2),
                    round(losses['V_ohmic_mV'][i], 2),
                    round(losses['V_mt_mV'][i], 2),
                ])

    # ── Sheet 5: Model Fit ──
    if fit_result is not None:
        ws5 = wb.create_sheet("Model Fit")

        # Fitted parameters
        ws5.append(['Fitted Parameters'])
        xf = fit_result['x']
        T_K = fit_result['T_K']
        params = [
            ('E_rev [V]', round(fit_result['E_rev'], 4)),
            ('ASR_total [mΩ·cm²]', round(xf[0], 1)),
            ('j0_anode [A/cm²]', f"{10**xf[1]:.3e}"),
            ('α_anode', round(xf[2], 3)),
            ('Anode Tafel slope [mV/dec]', round((_R*T_K)/(xf[2]*_n_e*_F)*1000, 1)),
            ('j0_cathode [A/cm²]', f"{10**xf[3]:.3e}"),
            ('Cathode Tafel slope [mV/dec]', round((_R*T_K)/(0.5*_n_e*_F)*1000, 1)),
            ('c_mt [V·cm⁴/A²]', round(xf[4], 5)),
            ('RMSE [mV]', round(fit_result['rmse_mV'], 2)),
            ('MAE [mV]', round(fit_result['mae_mV'], 2)),
            ('Max |error| [mV]', round(fit_result['max_err_mV'], 2)),
        ]
        for name, val in params:
            ws5.append([name, val])

        # Data vs model
        ws5.append([])
        ws5.append(['j [A/cm²]', 'V_data [V]', 'V_model [V]', 'Residual [mV]'])
        for j_val, v_dat, v_mod, res in zip(
                fit_result['j_data'], fit_result['V_data'],
                fit_result['V_model'], fit_result['residual']):
            ws5.append([round(j_val, 6), round(v_dat, 5), round(v_mod, 5),
                        round(res * 1000, 2)])

        # Model curve components
        ws5.append([])
        comp = fit_result['components']
        ws5.append(['j_model [A/cm²]', 'V_total [V]', 'E_rev [V]',
                    'η_anode [V]', 'η_cathode [V]', 'V_ohmic [V]', 'V_mt [V]'])
        # Export every 10th point of the smooth curve to keep file manageable
        step = max(1, len(comp['j']) // 50)
        for idx in range(0, len(comp['j']), step):
            ws5.append([
                round(comp['j'][idx], 6),
                round(comp['V_total'][idx], 5),
                round(comp['E_rev'][idx], 5),
                round(comp['eta_anode'][idx], 5),
                round(comp['eta_cathode'][idx], 5),
                round(comp['V_ohmic'][idx], 5),
                round(comp['V_mt'][idx], 5),
            ])

    # ── Sheet 6: EIS ──
    if eis_results:
        ws6 = wb.create_sheet("EIS")
        ws6.append(['File', 'Frequency [Hz]', "Z' [Ω]", "-Z'' [Ω]",
                     "Z' [mΩ·cm²]", "-Z'' [mΩ·cm²]",
                     'DC Voltage [V]', 'Elapsed Time [s]'])
        for er in eis_results:
            eis = er['eis_data']
            fname = Path(er['file']).name
            for i in range(len(eis['freq'])):
                row = [
                    fname,
                    round(eis['freq'][i], 3),
                    round(eis['zre'][i], 8),
                    round(eis['zim'][i], 8),
                    round(eis['zre'][i] * geo_area * 1000, 4),
                    round(eis['zim'][i] * geo_area * 1000, 4),
                ]
                if eis.get('dc_v') is not None:
                    row.append(round(eis['dc_v'][i], 5))
                else:
                    row.append('')
                if eis.get('time') is not None:
                    row.append(round(eis['time'][i], 2))
                else:
                    row.append('')
                ws6.append(row)

    # ── Sheet 7: iR Correction ──
    if ir_data is not None:
        ws7 = wb.create_sheet("iR Correction")
        ws7.append(['j [A/cm²]', 'V_raw [V]', 'V_iR-free [V]',
                    'ASR_interp [mΩ·cm²]', 'iR_drop [mV]'])
        for i in range(len(ir_data['j_pol'])):
            j_val = ir_data['j_pol'][i]
            ir_drop = j_val * ir_data['asr_interp'][i]  # mV
            ws7.append([
                round(j_val, 6),
                round(ir_data['V_pol'][i], 5),
                round(ir_data['V_irfree'][i], 5),
                round(ir_data['asr_interp'][i], 2),
                round(ir_drop, 2),
            ])
        # Add HFR measurement points
        ws7.append([])
        ws7.append(['HFR Measurements'])
        ws7.append(['j [A/cm²]', 'ASR [mΩ·cm²]'])
        for i in range(len(ir_data['j_hfr'])):
            ws7.append([round(ir_data['j_hfr'][i], 6),
                        round(ir_data['asr_hfr'][i], 2)])

    wb.save(filepath)
    print(f"  Data exported: {filepath}")


# ═══════════════════════════════════════════════════════════════════
#  EIS loading & HFR analysis
# ═══════════════════════════════════════════════════════════════════

def detect_eis_columns(fieldnames):
    """Auto-detect EIS columns: frequency, Z', -Z'', elapsed time, DC voltage."""
    result = {k: None for k in ('freq_col', 'zre_col', 'zim_col', 'time_col', 'dc_v_col')}

    freq_cands = ['frequency (hz)', 'freq (hz)', 'frequency(hz)', 'freq(hz)']
    zre_cands  = ["z' (ohms)", "z'(ohms)", "zre (ohms)", "z' (ohm)", "zreal"]
    zim_cands  = ['-z" (ohms)', '-z"(ohms)', "-z'' (ohms)", '-zim', '-z" (ohm)',
                  "z\" (ohms)", "z''"]
    time_cands = ['elapsed time', 'time (s)', 'elapsed_time', 'time(s)']
    dcv_cands  = ['dc working electrode (v)', 'working electrode (v)',
                  'dc voltage', 'ewe (v)', 'dc potential']

    for fn in fieldnames:
        fl = fn.lower().strip()

        if result['freq_col'] is None:
            for c in freq_cands:
                if c in fl:
                    result['freq_col'] = fn; break

        if result['zre_col'] is None:
            for c in zre_cands:
                if c in fl:
                    result['zre_col'] = fn; break

        if result['zim_col'] is None:
            for c in zim_cands:
                if c in fl:
                    result['zim_col'] = fn; break

        if result['time_col'] is None:
            for c in time_cands:
                if c in fl:
                    result['time_col'] = fn; break

        if result['dc_v_col'] is None:
            for c in dcv_cands:
                if c in fl:
                    result['dc_v_col'] = fn; break

    return result


def load_eis_data(filepath):
    """
    Load EIS data from CSV. Returns dict with arrays:
    freq, zre, zim, time (all sorted by descending frequency).
    """
    path = Path(_clean_path(filepath))
    if not path.exists():
        raise FileNotFoundError(f"EIS file not found: {path}")

    try:    text = path.read_text(encoding='utf-8')
    except: text = path.read_text(encoding='latin-1')

    delim = '\t' if '\t' in text.split('\n')[0] else ','
    lines = text.strip().split('\n')
    reader = csv.DictReader(lines, delimiter=delim)
    fn = [f.strip() for f in (reader.fieldnames or [])]
    reader.fieldnames = fn

    cols = detect_eis_columns(fn)
    if cols['freq_col'] is None or cols['zre_col'] is None or cols['zim_col'] is None:
        raise ValueError(
            f"Could not detect EIS columns.\n"
            f"  Available: {fn}\n"
            f"  Need: Frequency (Hz), Z' (Ohms), -Z\" (Ohms)")

    freq, zre, zim, time, dc_v = [], [], [], [], []
    for row in reader:
        try:
            freq.append(float(row[cols['freq_col']].strip()))
            zre.append(float(row[cols['zre_col']].strip()))
            zim.append(float(row[cols['zim_col']].strip()))
            if cols['time_col']:
                time.append(float(row[cols['time_col']].strip()))
            if cols['dc_v_col']:
                dc_v.append(float(row[cols['dc_v_col']].strip()))
        except (ValueError, KeyError):
            continue

    freq = np.array(freq)
    zre = np.array(zre)
    zim = np.array(zim)
    time = np.array(time) if time else None
    dc_v = np.array(dc_v) if dc_v else None

    # Sort by descending frequency
    order = np.argsort(freq)[::-1]
    freq, zre, zim = freq[order], zre[order], zim[order]
    if time is not None:
        time = time[order]
    if dc_v is not None:
        dc_v = dc_v[order]

    # Mean DC voltage during EIS measurement
    dc_v_mean = float(np.nanmean(dc_v)) if dc_v is not None else None

    print(f"  EIS: {len(freq)} points from '{path.name}'")
    print(f"    Columns: freq='{cols['freq_col']}', "
          f"Z'='{cols['zre_col']}', -Z''='{cols['zim_col']}'")
    if dc_v_mean is not None:
        print(f"    DC voltage: {dc_v_mean:.4f} V")
    print(f"    Freq range: {freq.min():.1f} – {freq.max():.0f} Hz")
    if time is not None:
        print(f"    Time range: {time.min():.1f} – {time.max():.1f} s")

    return {'freq': freq, 'zre': zre, 'zim': zim, 'time': time,
            'dc_v': dc_v, 'dc_v_mean': dc_v_mean, 'cols': cols}


def extract_hfr(eis, geo_area=5.0):
    """
    Extract HFR (high-frequency resistance) from EIS data.

    Finds the Z' intercept where -Z'' crosses zero from negative
    (inductive) to positive (capacitive). Uses linear interpolation
    between the two points bracketing the crossing.

    Returns
    -------
    hfr_ohm : float — HFR in Ohms
    asr_ohm_cm2 : float — ASR in Ω·cm²
    asr_mohm_cm2 : float — ASR in mΩ·cm²
    f_hfr : float — frequency at the crossing [Hz]
    """
    zre = eis['zre']
    zim = eis['zim']
    freq = eis['freq']

    # -Z'' convention: in the file, -Z'' is stored directly.
    # Inductive region: -Z'' < 0 (high frequency)
    # Capacitive region: -Z'' > 0 (lower frequency)
    # HFR = Z' where -Z'' crosses zero from negative to positive

    # Scan from high to low frequency for sign change in zim
    crossing_idx = None
    for i in range(len(zim) - 1):
        if zim[i] <= 0 and zim[i + 1] > 0:
            crossing_idx = i
            break

    if crossing_idx is not None:
        # Linear interpolation
        z1, z2 = zim[crossing_idx], zim[crossing_idx + 1]
        zr1, zr2 = zre[crossing_idx], zre[crossing_idx + 1]
        f1, f2 = freq[crossing_idx], freq[crossing_idx + 1]

        frac = -z1 / (z2 - z1)  # fraction from point 1 to point 2
        hfr = zr1 + frac * (zr2 - zr1)
        f_hfr = f1 + frac * (f2 - f1)
    else:
        # No sign change found — use minimum |Z''| point
        min_idx = np.argmin(np.abs(zim))
        hfr = zre[min_idx]
        f_hfr = freq[min_idx]

    asr = hfr * geo_area
    asr_m = asr * 1000

    print(f"\n  HFR Analysis:")
    print(f"    HFR           : {hfr*1000:.3f} mΩ  ({hfr:.6f} Ω)")
    print(f"    ASR           : {asr_m:.1f} mΩ·cm²  ({asr:.4f} Ω·cm²)")
    print(f"    Intercept freq: {f_hfr:.0f} Hz")

    return {'hfr_ohm': hfr, 'asr_ohm_cm2': asr, 'asr_mohm_cm2': asr_m,
            'f_hfr': f_hfr}


def plot_nyquist(eis, hfr_result, geo_area=5.0, save_path=None):
    """Nyquist plot with HFR intercept marked."""
    zre = eis['zre'] * geo_area * 1000   # → mΩ·cm²
    zim = eis['zim'] * geo_area * 1000

    fig, ax = plt.subplots(figsize=(7, 6), dpi=120)

    ax.plot(zre, zim, 'o-', ms=4, lw=1.2, color='#1f77b4')

    # Mark HFR intercept
    hfr_asr = hfr_result['asr_mohm_cm2']
    ax.plot(hfr_asr, 0, '*', ms=14, color='#d62728', markeredgecolor='k',
            markeredgewidth=0.8, zorder=5,
            label=f'HFR = {hfr_asr:.1f} mΩ·cm²')

    ax.axhline(0, color='k', lw=0.5, alpha=0.5)
    ax.set_xlabel("Z'  [mΩ·cm²]", fontsize=12)
    ax.set_ylabel("-Z''  [mΩ·cm²]", fontsize=12)
    ax.set_title('Nyquist Plot (EIS)', fontsize=12, fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.set_aspect('equal', adjustable='datalim')

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig





# ═══════════════════════════════════════════════════════════════════
#  Folder scanning & multi-EIS processing
# ═══════════════════════════════════════════════════════════════════

def scan_folder(folder_path, cell_id='a1'):
    """
    Scan a folder to find polcurve and EIS files for a given cell ID.

    Convention:
      - Polcurve file: starts with "1_" + cell_id (e.g. "1_a1_...")
      - EIS files: contain cell_id but don't start with "1_" + cell_id

    Returns
    -------
    polcurve_file : str or None
    eis_files : list of str (sorted by filename)
    """
    import glob, os
    folder = Path(_clean_path(folder_path))
    if not folder.is_dir():
        raise FileNotFoundError(f"Folder not found: {folder}")

    # Find all CSV files containing the cell ID
    all_csvs = sorted(glob.glob(str(folder / '*.csv')) +
                      glob.glob(str(folder / '*.CSV')))

    polcurve_file = None
    eis_files = []
    prefix = f'1_{cell_id}'

    for fp in all_csvs:
        base = os.path.basename(fp)
        if cell_id.lower() not in base.lower():
            continue

        if base.lower().startswith(prefix.lower()):
            polcurve_file = fp
        else:
            eis_files.append(fp)

    print(f"  Folder: {folder}")
    print(f"  Cell ID: '{cell_id}'")
    if polcurve_file:
        print(f"  Polcurve: {os.path.basename(polcurve_file)}")
    else:
        print(f"  WARNING: No polcurve file found (expected '1_{cell_id}_...')")
    print(f"  EIS files: {len(eis_files)}")
    for fp in eis_files:
        print(f"    {os.path.basename(fp)}")

    return polcurve_file, eis_files


def load_all_eis(eis_files, geo_area=5.0):
    """
    Load multiple EIS files, extract HFR from each.

    Returns list of dicts with HFR results + timestamp + DC voltage,
    sorted by time.
    """
    results = []

    for fp in eis_files:
        try:
            eis = load_eis_data(fp)
            hfr = extract_hfr(eis, geo_area=geo_area)

            t_eis = float(np.mean(eis['time'])) if eis['time'] is not None else None

            results.append({
                'file': fp,
                'hfr_ohm': hfr['hfr_ohm'],
                'asr_mohm_cm2': hfr['asr_mohm_cm2'],
                'f_hfr': hfr['f_hfr'],
                't_eis': t_eis,
                'dc_v_mean': eis.get('dc_v_mean'),
                'eis_data': eis,
            })
        except Exception as e:
            print(f"    WARNING: Failed to process {Path(fp).name}: {e}")

    # Sort by timestamp
    results.sort(key=lambda r: r['t_eis'] if r['t_eis'] is not None else 0)

    return results


def map_eis_to_cycles(eis_results, cycles):
    """
    Map each EIS measurement to the polcurve cycle that follows it.

    EIS characterizes the cell state before the next polcurve sweep,
    so each EIS is assigned to the first cycle whose start time is
    after the EIS measurement time.

    Returns list of dicts with cycle number and ASR.
    """
    # Compute start time for each cycle (earliest dwell midpoint)
    cycle_times = []
    for ci, cyc in enumerate(cycles):
        times = [d.get('t_mid') for d in cyc if d.get('t_mid') is not None]
        if times:
            cycle_times.append((ci + 1, min(times)))
        else:
            cycle_times.append((ci + 1, None))

    mapped = []
    for er in eis_results:
        t_eis = er['t_eis']
        if t_eis is None:
            continue

        # Find the first cycle that starts after the EIS measurement
        assigned_cyc = None
        for cn, ct in cycle_times:
            if ct is not None and ct > t_eis:
                assigned_cyc = cn
                break

        # If no cycle follows the EIS, assign to the last cycle
        if assigned_cyc is None:
            assigned_cyc = cycle_times[-1][0] if cycle_times else 0

        mapped.append({
            'cycle': assigned_cyc,
            'asr_mohm_cm2': er['asr_mohm_cm2'],
            't_eis': t_eis,
        })

    return mapped


def plot_j_and_hfr_vs_cycle(cycles, v_targets, eis_mapped, save_path=None):
    """
    Dual-axis plot: j at target voltages (left) and ASR from HFR (right)
    vs cycle number.
    """
    fig, ax1 = plt.subplots(figsize=(10, 5.5), dpi=120)
    ax2 = ax1.twinx()

    colors_j = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
    markers_j = ['o', 's', '^', 'D']

    # ── Left axis: j at target voltages ──
    stable_cycles = []
    for k, vt in enumerate(v_targets):
        cn, jv = extract_j_at_voltage(cycles, vt)
        if len(cn) == 0:
            continue
        c = colors_j[k % len(colors_j)]
        m = markers_j[k % len(markers_j)]
        ax1.plot(cn, jv, f'{m}-', color=c, ms=5, lw=1.2,
                 label=f'j @ {vt:.2f} V')

        sc = detect_stabilization(cn, jv)
        if sc is not None:
            stable_cycles.append(sc)
            sc_idx = np.where(cn == sc)[0]
            if len(sc_idx) > 0:
                ax1.plot(sc, jv[sc_idx[0]], '*', color=c, ms=14,
                         markeredgecolor='k', markeredgewidth=0.8, zorder=6)

    ax1.set_xlabel('Cycle number', fontsize=12)
    ax1.set_ylabel('Current density  j  [A/cm²]', fontsize=12, color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.set_xlim(left=0)
    ax1.grid(True, alpha=0.3)

    # ── Right axis: ASR from HFR ──
    if eis_mapped:
        cyc_nums = [em['cycle'] for em in eis_mapped]
        asr_vals = [em['asr_mohm_cm2'] for em in eis_mapped]
        ax2.plot(cyc_nums, asr_vals, 'v--', color='#9467bd', ms=8, lw=1.5,
                 markeredgecolor='k', markeredgewidth=0.5,
                 label='ASR (HFR)')
        ax2.set_ylabel('ASR  [mΩ·cm²]', fontsize=12, color='#9467bd')
        ax2.tick_params(axis='y', labelcolor='#9467bd')

    # Combined legend below x-axis
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(lines1 + lines2, labels1 + labels2,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(lines1) + len(lines2), fontsize=9,
               frameon=True, fancybox=True)

    ax1.set_title('Current Density & HFR vs. Cycle', fontsize=12, fontweight='bold')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig



# ═══════════════════════════════════════════════════════════════════
#  Current-dependent EIS & iR correction
# ═══════════════════════════════════════════════════════════════════

def detect_current_dependent_eis(eis_results, cycles, v_tol=0.03):
    """
    Detect if EIS measurements were taken at multiple voltages during
    a polcurve cycle (current-dependent EIS).

    Returns
    -------
    is_current_dep : bool
    eis_at_j : list of dicts with 'j', 'V', 'asr_mohm_cm2', 'hfr_ohm'
        EIS measurements matched to polcurve j values, sorted by j
    cycle_idx : int or None
        Index of the polcurve cycle the EIS corresponds to
    """
    if not eis_results or not cycles:
        return False, [], None

    # Get unique DC voltages across all EIS files
    dc_voltages = [er['dc_v_mean'] for er in eis_results
                   if er['dc_v_mean'] is not None]
    if len(dc_voltages) < 2:
        return False, [], None

    # Check if they span a significant voltage range (> 100 mV)
    v_span = max(dc_voltages) - min(dc_voltages)
    if v_span < 0.10:
        return False, [], None

    # Find which polcurve cycle these EIS are closest to
    # Use the first EIS measurement's time to find the cycle
    t_eis_mean = np.mean([er['t_eis'] for er in eis_results
                          if er['t_eis'] is not None])

    # Find cycle with start time closest after the EIS block
    best_cyc_idx = None
    for ci, cyc in enumerate(cycles):
        times = [d.get('t_mid') for d in cyc if d.get('t_mid') is not None]
        if times and min(times) > t_eis_mean:
            best_cyc_idx = ci
            break
    if best_cyc_idx is None:
        best_cyc_idx = len(cycles) - 1

    ref_cyc = cycles[best_cyc_idx]

    # Match each EIS to a polcurve j by DC voltage
    eis_at_j = []
    for er in eis_results:
        v_eis = er['dc_v_mean']
        if v_eis is None:
            continue

        # Find the polcurve dwell closest in voltage
        best_d = None
        best_dv = v_tol + 1
        for d in ref_cyc:
            dv = abs(d['V'] - v_eis)
            if dv < best_dv:
                best_dv = dv
                best_d = d

        if best_d is not None and best_dv <= v_tol:
            eis_at_j.append({
                'j': best_d['j'],
                'V': best_d['V'],
                'asr_mohm_cm2': er['asr_mohm_cm2'],
                'hfr_ohm': er['hfr_ohm'],
                'dc_v_eis': v_eis,
            })

    # Sort by j
    eis_at_j.sort(key=lambda x: x['j'])

    is_current_dep = len(eis_at_j) >= 3
    if is_current_dep:
        print(f"\n  Current-dependent EIS detected ({len(eis_at_j)} points):")
        for e in eis_at_j:
            print(f"    V = {e['V']:.3f} V, j = {e['j']:.3f} A/cm², "
                  f"ASR = {e['asr_mohm_cm2']:.1f} mΩ·cm²")

    return is_current_dep, eis_at_j, best_cyc_idx


def compute_ir_corrected_polcurve(j_pol, V_pol, eis_at_j, geo_area=5.0):
    """
    Compute iR-corrected polarization curve using current-dependent HFR.

    Interpolates HFR(j) from the EIS measurements, then computes:
      V_iR-free(j) = V(j) - j × HFR(j)

    Returns
    -------
    j_pol : array — current density [A/cm²]
    V_pol : array — raw cell voltage [V]
    V_irfree : array — iR-corrected voltage [V]
    j_hfr : array — j values where HFR was measured
    asr_hfr : array — ASR at each j [mΩ·cm²]
    asr_interp : array — interpolated ASR at each polcurve j [mΩ·cm²]
    """
    j_hfr = np.array([e['j'] for e in eis_at_j])
    asr_hfr = np.array([e['asr_mohm_cm2'] for e in eis_at_j])

    # Interpolate ASR to all polcurve j values
    # Clip to the range of measured HFR (no extrapolation)
    asr_interp = np.interp(j_pol, j_hfr, asr_hfr)

    # iR correction: V_irfree = V - j × ASR (convert mΩ·cm² to Ω·cm²)
    V_irfree = V_pol - j_pol * asr_interp / 1000.0

    return j_pol, V_pol, V_irfree, j_hfr, asr_hfr, asr_interp


def plot_ir_correction(j_pol, V_pol, V_irfree, j_hfr, asr_hfr, asr_interp,
                       cycle_label=None, save_path=None):
    """
    Three-panel plot:
      Left  — Raw and iR-corrected polcurves
      Center — ASR (HFR) vs current density
      Right — iR drop vs current density
    """
    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(16, 5.5), dpi=120)
    ttl = cycle_label or 'Current-Dependent EIS Analysis'
    fig.suptitle(ttl, fontsize=12, fontweight='bold')

    # ── Left: polcurves ──
    ax1.plot(j_pol, V_pol, 'ko-', ms=5, lw=1.5, label='Raw V(j)')
    ax1.plot(j_pol, V_irfree, 's-', color='#2196F3', ms=5, lw=1.5,
             label='iR-corrected')
    ax1.set_xlabel('j  [A/cm²]', fontsize=11)
    ax1.set_ylabel('Cell voltage  [V]', fontsize=11)
    ax1.set_xlim(left=0)
    ax1.legend(fontsize=9)
    ax1.grid(True, alpha=0.3)
    ax1.set_title('Polarization Curve', fontsize=10)

    # ── Center: ASR vs j ──
    ax2.plot(j_hfr, asr_hfr, 'o', ms=7, color='#d62728',
             markeredgecolor='k', markeredgewidth=0.5, label='HFR (measured)')
    j_smooth = np.linspace(j_pol.min(), j_pol.max(), 200)
    asr_smooth = np.interp(j_smooth, j_hfr, asr_hfr)
    ax2.plot(j_smooth, asr_smooth, '--', color='#d62728', lw=1, alpha=0.6,
             label='Interpolated')
    ax2.set_xlabel('j  [A/cm²]', fontsize=11)
    ax2.set_ylabel('ASR  [mΩ·cm²]', fontsize=11)
    ax2.set_xlim(left=0)
    ax2.legend(fontsize=9)
    ax2.grid(True, alpha=0.3)
    ax2.set_title('HFR vs. Current Density', fontsize=10)

    # ── Right: iR drop ──
    ir_drop = j_pol * asr_interp  # mV
    ax3.plot(j_pol, ir_drop, 'o-', ms=5, color='#4CAF50', lw=1.5)
    ax3.set_xlabel('j  [A/cm²]', fontsize=11)
    ax3.set_ylabel('iR drop  [mV]', fontsize=11)
    ax3.set_xlim(left=0)
    ax3.grid(True, alpha=0.3)
    ax3.set_title('Ohmic Loss', fontsize=10)

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


# ═══════════════════════════════════════════════════════════════════
#  Electrolyzer model & fitting
# ═══════════════════════════════════════════════════════════════════

# Physical constants
_F   = 96485.3329   # C/mol
_R   = 8.31446      # J/(mol·K)
_n_e = 2            # electrons per H2O split


def E_rev(T_C=80.0, p_cathode_barg=0.0, p_anode_barg=0.0):
    """Nernst reversible potential (LeRoy & Bowen 1980 + pressure correction)."""
    T = T_C + 273.15
    E0 = 1.5184 - 1.5421e-3*T + 9.523e-5*T*np.log(T) + 9.84e-8*T**2
    p_H2 = p_cathode_barg + 1.01325
    p_O2 = p_anode_barg + 1.01325
    nernst = (_R * T) / (_n_e * _F) * np.log(p_H2 * p_O2**0.5)
    return E0 + nernst


def _electrolyzer_model(j, x, E, T_K):
    """V(j) = E_rev + η_a + η_c + j·ASR + c_mt·j²"""
    ASR, log_j0a, alpha_a, log_j0c, c_mt = x
    j0a, j0c = 10**log_j0a, 10**log_j0c
    ba = (_R * T_K) / (alpha_a * _n_e * _F)
    bc = (_R * T_K) / (0.5 * _n_e * _F)
    eta_a = np.where(j > j0a, ba * np.log10(j / j0a), 0.0)
    eta_c = np.where(j > j0c, bc * np.log10(j / j0c), 0.0)
    return E + eta_a + eta_c + j * ASR / 1000.0 + c_mt * j**2


def fit_polcurve(j_data, V_data, T_C=80.0, p_cathode_barg=0.0,
                 p_anode_barg=0.0, fix_ASR=None):
    """
    Fit electrolyzer model to a single polcurve.

    Model: V = E_rev(T,p) + η_a(Tafel) + η_c(Tafel) + j·ASR + c_mt·j²

    Parameters
    ----------
    j_data, V_data : arrays of current density [A/cm²] and voltage [V]
    T_C            : cell temperature [°C]
    p_cathode_barg : cathode H2 pressure [barg]
    p_anode_barg   : anode O2 pressure [barg]
    fix_ASR        : if set, pin ASR to this value [mΩ·cm²]

    Returns
    -------
    result : dict with fitted parameters, model curve, residuals, stats
    """
    from scipy.optimize import least_squares

    E = E_rev(T_C, p_cathode_barg, p_anode_barg)
    T_K = T_C + 273.15

    # Filter j > 0
    mask = j_data > 0
    j_fit = j_data[mask]
    V_fit = V_data[mask]

    def model(j, x):
        return _electrolyzer_model(j, x, E, T_K)

    def residuals(x):
        return model(j_fit, x) - V_fit

    # x = [ASR_mOhm, log10(j0_a), alpha_a, log10(j0_c), c_mt]
    x0 = [70.0, -7.0, 0.5, -3.0, 0.0]
    lo = [10.0, -12.0, 0.2, -6.0, 0.0]
    hi = [500.0, -3.0, 2.0, -0.5, 0.05]

    if fix_ASR is not None:
        x0[0], lo[0], hi[0] = fix_ASR, fix_ASR - 0.01, fix_ASR + 0.01

    res = least_squares(residuals, x0, bounds=(lo, hi), method='trf',
                        loss='soft_l1', f_scale=0.01)

    xf = res.x
    V_model = model(j_fit, xf)
    rd = V_fit - V_model
    rmse = np.sqrt(np.mean(rd**2))

    # Compute individual loss components over a smooth j range
    j_smooth = np.linspace(0, j_fit.max() * 1.1, 500)
    j0a, j0c = 10**xf[1], 10**xf[3]
    ba = (_R * T_K) / (xf[2] * _n_e * _F)
    bc = (_R * T_K) / (0.5 * _n_e * _F)

    components = {
        'j': j_smooth,
        'E_rev': np.full_like(j_smooth, E),
        'eta_anode': np.where(j_smooth > j0a, ba * np.log10(j_smooth / j0a), 0.0),
        'eta_cathode': np.where(j_smooth > j0c, bc * np.log10(j_smooth / j0c), 0.0),
        'V_ohmic': j_smooth * xf[0] / 1000.0,
        'V_mt': xf[4] * j_smooth**2,
    }
    components['V_total'] = (components['E_rev'] + components['eta_anode'] +
                              components['eta_cathode'] + components['V_ohmic'] +
                              components['V_mt'])

    return {
        'x': xf, 'E_rev': E, 'T_C': T_C, 'T_K': T_K,
        'p_cathode_barg': p_cathode_barg, 'p_anode_barg': p_anode_barg,
        'j_data': j_fit, 'V_data': V_fit,
        'V_model': V_model, 'residual': rd,
        'rmse_mV': rmse * 1000,
        'mae_mV': np.mean(np.abs(rd)) * 1000,
        'max_err_mV': np.max(np.abs(rd)) * 1000,
        'components': components,
        'converged': res.success,
        'message': res.message,
    }


def print_fit_summary(fr):
    """Print fitted model parameters and loss breakdown at key current densities."""
    xf = fr['x']
    T_K = fr['T_K']

    print("=" * 60)
    print("  Electrolyzer Model Fit — Last Cycle")
    print("=" * 60)
    print(f"  Temperature          : {fr['T_C']:.1f} °C")
    print(f"  Cathode pressure     : {fr['p_cathode_barg']:.1f} barg")
    print(f"  Anode pressure       : {fr['p_anode_barg']:.1f} barg")
    print(f"  E_rev (fixed)        : {fr['E_rev']:.4f} V")
    print()
    print(f"  ASR_total            : {xf[0]:.1f} mΩ·cm²")
    print(f"  j0_anode             : {10**xf[1]:.3e} A/cm²")
    print(f"  α_anode              : {xf[2]:.3f}")
    print(f"  Anode Tafel slope    : {(_R*T_K)/(xf[2]*_n_e*_F)*1000:.1f} mV/dec")
    print(f"  j0_cathode           : {10**xf[3]:.3e} A/cm²")
    print(f"  Cathode Tafel slope  : {(_R*T_K)/(0.5*_n_e*_F)*1000:.1f} mV/dec (α_c = 0.5)")
    print(f"  c_mt                 : {xf[4]:.5f} V·cm⁴/A²")
    print()
    print(f"  RMSE                 : {fr['rmse_mV']:.2f} mV")
    print(f"  MAE                  : {fr['mae_mV']:.2f} mV")
    print(f"  Max |error|          : {fr['max_err_mV']:.2f} mV")
    print(f"  Converged            : {fr['converged']}")
    print()

    # Loss breakdown at key current densities
    E = fr['E_rev']
    V_tn = 1.481
    j0a, j0c = 10**xf[1], 10**xf[3]
    ba = (_R * T_K) / (xf[2] * _n_e * _F) * 1000  # mV/dec
    bc = (_R * T_K) / (0.5 * _n_e * _F) * 1000

    print(f"  {'j [A/cm²]':>12} {'V [V]':>10} {'η_a [mV]':>10} {'η_c [mV]':>10} "
          f"{'Ohm [mV]':>10} {'MT [mV]':>10} {'η_HHV [%]':>10}")
    print("  " + "-" * 75)

    for jj in [0.1, 0.5, 1.0, 2.0, 3.0, 4.0, 5.0]:
        if jj > fr['j_data'].max() * 1.2:
            continue
        j_arr = np.array([jj])
        V = _electrolyzer_model(j_arr, xf, E, T_K)[0]
        ea = ba/1000 * np.log10(jj / j0a) if jj > j0a else 0
        ec = bc/1000 * np.log10(jj / j0c) if jj > j0c else 0
        vo = jj * xf[0] / 1000
        vm = xf[4] * jj**2
        eff = V_tn / V * 100 if V > 0 else 0
        print(f"  {jj:12.2f} {V:10.4f} {ea*1000:10.1f} {ec*1000:10.1f} "
              f"{vo*1000:10.1f} {vm*1000:10.1f} {eff:10.1f}")
    print("=" * 60)
    print()


def plot_fit(fr, save_path=None):
    """
    Three-panel plot: stacked-area loss breakdown, data vs model, residuals.
    """
    comp = fr['components']
    j_m = comp['j']
    jd, Vd = fr['j_data'], fr['V_data']

    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(18, 5.5), dpi=120)
    fig.suptitle(
        f"Electrolyzer Model Fit — Last Cycle\n"
        f"RMSE = {fr['rmse_mV']:.1f} mV,  ASR = {fr['x'][0]:.1f} mΩ·cm²,  "
        f"T = {fr['T_C']:.0f} °C",
        fontsize=11, fontweight='bold')

    # Left: stacked area
    labels = ['E_rev', 'η anode (OER)', 'η cathode (HER)', 'Ohmic', 'Mass transport']
    colors = ['#2196F3', '#FF5722', '#FF9800', '#4CAF50', '#9C27B0']
    arrays = [comp['E_rev'], comp['eta_anode'], comp['eta_cathode'],
              comp['V_ohmic'], comp['V_mt']]
    ax1.stackplot(j_m, *arrays, labels=labels, colors=colors, alpha=0.7)
    ax1.plot(jd, Vd, 'ko', ms=5, label='Data', zorder=5)
    ax1.set_xlabel('j  [A/cm²]'); ax1.set_ylabel('V  [V]')
    ax1.set_xlim(0, j_m.max()); ax1.legend(loc='upper left', fontsize=7)
    ax1.grid(True, alpha=0.3); ax1.set_title('Loss breakdown', fontsize=10)

    # Center: data vs model
    ax2.plot(jd, Vd, 'ko', ms=5, label='Data', zorder=5)
    ax2.plot(j_m, comp['V_total'], 'r-', lw=2, label='Fitted model')
    ax2.set_xlabel('j  [A/cm²]'); ax2.set_ylabel('V  [V]')
    ax2.set_xlim(0, j_m.max()); ax2.legend(loc='upper left', fontsize=9)
    ax2.grid(True, alpha=0.3); ax2.set_title('Data vs. model', fontsize=10)

    # Right: residuals
    rm = fr['residual'] * 1000
    ax3.stem(jd, rm, linefmt='C0-', markerfmt='C0o', basefmt='k-')
    ax3.axhline(0, color='k', lw=0.5)
    ax3.axhspan(-fr['rmse_mV'], fr['rmse_mV'], alpha=0.15, color='green',
                label=f'±RMSE ({fr["rmse_mV"]:.1f} mV)')
    ax3.set_xlabel('j  [A/cm²]'); ax3.set_ylabel('Residual [mV]')
    ax3.legend(fontsize=9); ax3.grid(True, alpha=0.3)
    ax3.set_title('Residuals', fontsize=10)

    plt.tight_layout()
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig



# ═══════════════════════════════════════════════════════════════════
#  Per-cycle loss tracking
# ═══════════════════════════════════════════════════════════════════

def compute_losses_at_j(j_val, xf, T_K):
    """
    Compute individual voltage losses at a single current density
    from fitted parameters.

    Returns dict with eta_anode, eta_cathode, V_ohmic, V_mt in mV.
    """
    ASR, log_j0a, alpha_a, log_j0c, c_mt = xf
    j0a, j0c = 10**log_j0a, 10**log_j0c
    ba = (_R * T_K) / (alpha_a * _n_e * _F)
    bc = (_R * T_K) / (0.5 * _n_e * _F)

    eta_a = ba * np.log10(j_val / j0a) * 1000 if j_val > j0a else 0.0
    eta_c = bc * np.log10(j_val / j0c) * 1000 if j_val > j0c else 0.0
    v_ohm = j_val * ASR  # ASR in mΩ·cm², j in A/cm² → mV
    v_mt  = c_mt * j_val**2 * 1000  # V → mV

    return {'eta_anode_mV': eta_a, 'eta_cathode_mV': eta_c,
            'V_ohmic_mV': v_ohm, 'V_mt_mV': v_mt}


def extract_losses_vs_cycle(cycles, v_target, T_C=80.0,
                            p_cathode_barg=0.0, p_anode_barg=0.0,
                            fix_ASR=None, v_tol=0.015):
    """
    Fit each complete cycle and extract j + loss breakdown at v_target.

    Returns
    -------
    cycle_nums : list of cycle numbers
    j_values   : list of j at v_target [A/cm²]
    losses     : dict of loss name → list of values [mV]
    """
    from scipy.optimize import least_squares

    T_K = T_C + 273.15
    E = E_rev(T_C, p_cathode_barg, p_anode_barg)

    # Determine which cycles are "complete" (max setpoint count)
    max_pts = max(len(c) for c in cycles) if cycles else 0
    min_pts = max(5, int(max_pts * 0.8))

    cycle_nums = []
    j_values = []
    losses = {'eta_anode_mV': [], 'eta_cathode_mV': [],
              'V_ohmic_mV': [], 'V_mt_mV': []}

    print(f"\n  Fitting {sum(1 for c in cycles if len(c) >= min_pts)} cycles "
          f"for loss tracking at {v_target:.2f} V...")

    for ci, cyc in enumerate(cycles):
        if len(cyc) < min_pts:
            continue

        j_arr = np.array([d['j'] for d in cyc])
        V_arr = np.array([d['V'] for d in cyc])

        # Find j at target voltage
        best_j, best_dv = None, v_tol + 1
        for d in cyc:
            dv = abs(d['V'] - v_target)
            if dv < best_dv:
                best_dv = dv
                best_j = d['j']
        if best_dv > v_tol or best_j is None:
            continue

        # Fit this cycle (silent)
        mask = j_arr > 0
        j_fit, V_fit = j_arr[mask], V_arr[mask]
        if len(j_fit) < 4:
            continue

        def model(j, x):
            return _electrolyzer_model(j, x, E, T_K)

        x0 = [70.0, -7.0, 0.5, -3.0, 0.0]
        lo = [10.0, -12.0, 0.2, -6.0, 0.0]
        hi = [500.0, -3.0, 2.0, -0.5, 0.05]
        if fix_ASR is not None:
            x0[0], lo[0], hi[0] = fix_ASR, fix_ASR - 0.01, fix_ASR + 0.01

        try:
            res = least_squares(lambda x: model(j_fit, x) - V_fit,
                                x0, bounds=(lo, hi), method='trf',
                                loss='soft_l1', f_scale=0.01)
            if not res.success:
                continue
        except Exception:
            continue

        # Compute losses at the target j
        loss = compute_losses_at_j(best_j, res.x, T_K)

        cycle_nums.append(ci + 1)
        j_values.append(best_j)
        for k in losses:
            losses[k].append(loss[k])

    # Convert to arrays
    cycle_nums = np.array(cycle_nums)
    j_values = np.array(j_values)
    losses = {k: np.array(v) for k, v in losses.items()}

    print(f"    {len(cycle_nums)} cycles fitted successfully")

    return cycle_nums, j_values, losses


def plot_j_and_losses_vs_cycle(cycle_nums, j_values, losses,
                                v_target, save_path=None):
    """
    Dual-axis plot: j at target voltage (left) and voltage losses (right)
    vs cycle number.
    """
    fig, ax1 = plt.subplots(figsize=(10, 6), dpi=120)
    ax2 = ax1.twinx()

    # ── Left axis: j ──
    ax1.plot(cycle_nums, j_values, 'o-', color='#1f77b4', ms=5, lw=1.5,
             label=f'j @ {v_target:.2f} V')
    ax1.set_xlabel('Cycle number', fontsize=12)
    ax1.set_ylabel(f'Current density at {v_target:.2f} V  [A/cm²]',
                   fontsize=12, color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.set_xlim(left=0)

    # ── Right axis: losses ──
    loss_styles = [
        ('V_ohmic_mV',      'Ohmic',            '#4CAF50', 's-'),
        ('eta_anode_mV',    'η anode (OER)',     '#FF5722', '^-'),
        ('eta_cathode_mV',  'η cathode (HER)',   '#FF9800', 'v-'),
        ('V_mt_mV',         'Mass transport',    '#9C27B0', 'D-'),
    ]

    for key, label, color, fmt in loss_styles:
        vals = losses[key]
        if np.max(vals) < 0.1:  # skip if negligible
            continue
        ax2.plot(cycle_nums, vals, fmt, color=color, ms=4, lw=1.2,
                 label=label, alpha=0.85)

    # Total kinetic losses (anode + cathode)
    eta_total_kinetic = losses['eta_anode_mV'] + losses['eta_cathode_mV']
    ax2.plot(cycle_nums, eta_total_kinetic, 'p-', color='#d62728', ms=5, lw=1.5,
             label='η kinetic (total)', alpha=0.85)

    ax2.set_ylabel('Voltage loss  [mV]', fontsize=12)
    ax2.tick_params(axis='y')

    # Combined legend below x-axis
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(lines1 + lines2, labels1 + labels2,
               loc='lower center', bbox_to_anchor=(0.5, -0.02),
               ncol=len(lines1) + len(lines2), fontsize=9,
               frameon=True, fancybox=True)

    ax1.grid(True, alpha=0.3)
    ax1.set_title(f'Current Density & Losses at {v_target:.2f} V vs. Cycle',
                  fontsize=12, fontweight='bold')

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    if save_path:
        fig.savefig(save_path, bbox_inches='tight')
        print(f"  Plot saved: {save_path}")
    else:
        plt.show()
    return fig


# ═══════════════════════════════════════════════════════════════════
#  Main pipeline
# ═══════════════════════════════════════════════════════════════════

def analyze(filepath, geo_area=5.0, save_dir=None, title=None,
            T_C=80.0, p_cathode_barg=0.0, p_anode_barg=0.0,
            eis_files=None, eis_ref_voltage=None):
    """Full pipeline: load → extract → cycles → EIS → plot → fit."""

    # Load polcurve
    data, fieldnames = load_data(filepath)
    cols = detect_columns(fieldnames)

    if cols['v_col'] is None:
        raise ValueError(f"Could not auto-detect voltage column.\n  Available: {fieldnames}")
    if cols['i_col'] is None:
        raise ValueError(f"Could not auto-detect current column.\n  Available: {fieldnames}")

    print(f"\n  Detected columns:")
    print(f"    Voltage : '{cols['v_col']}'")
    print(f"    Current : '{cols['i_col']}'")
    if cols['step_col']:   print(f"    Step    : '{cols['step_col']}'")
    if cols['repeat_col']: print(f"    Repeat  : '{cols['repeat_col']}'")
    if cols['time_col']:   print(f"    Time    : '{cols['time_col']}'")
    print(f"\n  Cell area : {geo_area:.2f} cm²")

    # Extract dwells
    if cols['step_col'] and cols['repeat_col']:
        dwells = extract_dwells_from_steps(data, cols, geo_area)
        print(f"  Dwells extracted: {len(dwells)} (from step/repeat structure)")
    else:
        dwells = extract_dwells_generic(data, cols, geo_area)
        print(f"  Dwells extracted: {len(dwells)} (from voltage grouping)")

    if not dwells:
        print("  ERROR: No valid dwells found.")
        return None, None, None

    # Free raw data — no longer needed after dwell extraction
    del data
    gc.collect()

    # Detect cycles
    cycles = detect_cycles(dwells)
    print(f"  Cycles detected: {len(cycles)}")
    for i, cyc in enumerate(cycles):
        V_lo, V_hi = min(d['V'] for d in cyc), max(d['V'] for d in cyc)
        j_lo, j_hi = min(d['j'] for d in cyc), max(d['j'] for d in cyc)
        print(f"    Cycle {i+1:3d}: {len(cyc):2d} pts, "
              f"V = {V_lo:.3f}–{V_hi:.3f} V, j = {j_lo:.3f}–{j_hi:.3f} A/cm²")

    # Resolve output paths
    import os
    if save_dir:
        os.makedirs(save_dir, exist_ok=True)
        polcurve_path = os.path.join(save_dir, 'polcurve.png')
        jvc_path = os.path.join(save_dir, 'j_vs_cycle.png')
        xlsx_path = os.path.join(save_dir, 'analysis_data.xlsx')
        fit_path = os.path.join(save_dir, 'model_fit.png')
        nyquist_path = os.path.join(save_dir, 'nyquist.png')
        losses_path = os.path.join(save_dir, 'losses_vs_cycle.png')
        ir_path = os.path.join(save_dir, 'ir_correction.png')
    else:
        polcurve_path = jvc_path = xlsx_path = fit_path = nyquist_path = losses_path = ir_path = None

    # Plot polcurves
    plot_cycles(cycles, geo_area, title=title, save_path=polcurve_path)
    plt.close('all')
    gc.collect()

    # ── EIS / HFR analysis ──
    # Normalize eis_files to a list
    if eis_files is None:
        eis_files_list = []
    elif isinstance(eis_files, str):
        eis_files_list = [eis_files]
    else:
        eis_files_list = list(eis_files)

    eis_mapped = []
    fix_ASR = None
    eis_results_for_export = []

    if eis_files_list:
        print(f"\n  Loading {len(eis_files_list)} EIS file(s)...")
        eis_results = load_all_eis(eis_files_list, geo_area=geo_area)

        if eis_results:
            # ── Filter EIS by reference voltage ──
            if eis_ref_voltage is not None:
                v_tol = 0.02 * abs(eis_ref_voltage)  # 2% tolerance
                matched = [er for er in eis_results
                           if er['dc_v_mean'] is not None and
                           abs(er['dc_v_mean'] - eis_ref_voltage) <= v_tol]
                excluded = len(eis_results) - len(matched)
                print(f"\n  EIS voltage filter: ref = {eis_ref_voltage:.4f} V "
                      f"(±{v_tol*1000:.1f} mV, 2%)")
                print(f"    Matched : {len(matched)} file(s)")
                if excluded > 0:
                    matched_ids = set(id(m) for m in matched)
                    for er in eis_results:
                        if id(er) not in matched_ids:
                            v_str = f"{er['dc_v_mean']:.4f}" if er['dc_v_mean'] else "unknown"
                            print(f"    Excluded: {Path(er['file']).name} "
                                  f"(V = {v_str} V)")
                eis_results_for_tracking = matched
            else:
                eis_results_for_tracking = eis_results

            # Map filtered EIS to cycles by elapsed time
            eis_mapped = map_eis_to_cycles(eis_results_for_tracking, cycles)

            if eis_mapped:
                print(f"\n  EIS-to-cycle mapping:")
                for em in eis_mapped:
                    lbl = f"cycle {em['cycle']}" if em['cycle'] > 0 else "pre-conditioning"
                    print(f"    t = {em['t_eis']:.0f} s → {lbl}, "
                          f"ASR = {em['asr_mohm_cm2']:.1f} mΩ·cm²")

            # Use the last matched EIS measurement's ASR for model fit
            if eis_results_for_tracking:
                fix_ASR = eis_results_for_tracking[-1]['asr_mohm_cm2']

                # Plot Nyquist for last matched EIS
                last_eis = eis_results_for_tracking[-1]
                plot_nyquist(last_eis['eis_data'],
                             {'asr_mohm_cm2': fix_ASR,
                              'hfr_ohm': last_eis['hfr_ohm']},
                             geo_area=geo_area, save_path=nyquist_path)

            eis_results_for_export = eis_results_for_tracking
            plt.close('all')

    # ── Current-dependent EIS → iR correction ──
    ir_data = None
    if eis_files_list and eis_results:
        is_cd, eis_at_j, cd_cyc_idx = detect_current_dependent_eis(
            eis_results, cycles)
        if is_cd and eis_at_j:
            # Get the polcurve data for the matched cycle
            ref_cyc = cycles[cd_cyc_idx]
            j_pol = np.array([d['j'] for d in ref_cyc])
            V_pol = np.array([d['V'] for d in ref_cyc])
            # Only use polcurve points within the EIS j range
            mask = j_pol > 0
            j_pol, V_pol = j_pol[mask], V_pol[mask]

            j_p, V_p, V_irf, j_h, asr_h, asr_i = compute_ir_corrected_polcurve(
                j_pol, V_pol, eis_at_j, geo_area=geo_area)

            ir_data = {'j_pol': j_p, 'V_pol': V_p, 'V_irfree': V_irf,
                       'j_hfr': j_h, 'asr_hfr': asr_h, 'asr_interp': asr_i}

            plot_ir_correction(
                j_p, V_p, V_irf, j_h, asr_h, asr_i,
                cycle_label=f'iR Correction — Cycle {cd_cyc_idx + 1}',
                save_path=ir_path)
            plt.close('all')

    # ── Plot j and HFR vs cycle ──
    if len(cycles) >= 2:
        if eis_mapped:
            # Dual-axis: j + HFR
            plot_j_and_hfr_vs_cycle(cycles, [1.8, 1.7], eis_mapped,
                                     save_path=jvc_path)
        else:
            # j only
            plot_j_vs_cycle(cycles, [1.8, 1.7], save_path=jvc_path)
        plt.close('all')

    # ── Loss breakdown vs cycle at 1.8 V ──
    loss_data = None
    if len(cycles) >= 3:
        cn_loss, j_loss, losses = extract_losses_vs_cycle(
            cycles, v_target=1.8, T_C=T_C,
            p_cathode_barg=p_cathode_barg, p_anode_barg=p_anode_barg,
            fix_ASR=fix_ASR)
        if len(cn_loss) >= 2:
            loss_data = (cn_loss, j_loss, losses)
            plot_j_and_losses_vs_cycle(cn_loss, j_loss, losses,
                                       v_target=1.8, save_path=losses_path)
            plt.close('all')

    # ── Fit last complete cycle ──
    fr = None
    if cycles:
        max_pts = max(len(c) for c in cycles)
        full_cycles = [c for c in cycles if len(c) == max_pts]
        if not full_cycles:
            full_cycles = [c for c in cycles if len(c) >= max_pts * 0.8]

        if full_cycles:
            last_cyc = full_cycles[-1]
            cyc_num = cycles.index(last_cyc) + 1
            j_last = np.array([d['j'] for d in last_cyc])
            V_last = np.array([d['V'] for d in last_cyc])

            if fix_ASR:
                print(f"\n  Fitting cycle {cyc_num} ({len(last_cyc)} pts)"
                      f" with ASR fixed from HFR = {fix_ASR:.1f} mΩ·cm²...")
            else:
                print(f"\n  Fitting cycle {cyc_num} ({len(last_cyc)} pts)...")

            fr = fit_polcurve(j_last, V_last, T_C=T_C,
                              p_cathode_barg=p_cathode_barg,
                              p_anode_barg=p_anode_barg,
                              fix_ASR=fix_ASR)
            print_fit_summary(fr)
            plot_fit(fr, save_path=fit_path)
            plt.close('all')

    # ── Export Excel ──
    if xlsx_path:
        export_excel(xlsx_path, cycles, v_targets=[1.8, 1.7],
                     eis_mapped=eis_mapped if eis_mapped else None,
                     loss_data=loss_data, fit_result=fr,
                     eis_results=eis_results_for_export if eis_results_for_export else None,
                     ir_data=ir_data,
                     geo_area=geo_area)

    return cycles, fr, eis_mapped


# ═══════════════════════════════════════════════════════════════════
#  CLI & Interactive
# ═══════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description='PEM Electrolyzer Polcurve Analyzer')
    ap.add_argument('--file', type=str, default=None, help='Path to polcurve CSV')
    ap.add_argument('--eis', type=str, nargs='*', default=None,
                    help='Path(s) to EIS CSV file(s)')
    ap.add_argument('--folder', type=str, default=None,
                    help='Folder containing polcurve (1_a1_...) and EIS (*a1*) files')
    ap.add_argument('--cell-id', type=str, default='a1',
                    help='Cell identifier for folder scanning (default: a1)')
    ap.add_argument('--eis-ref-voltage', type=float, default=None,
                    help='Reference DC voltage for EIS filtering [V] (only include '
                         'EIS measured at this voltage ±2%%)')
    ap.add_argument('--area', type=float, default=None,
                    help='Geometric electrode area [cm²]')
    ap.add_argument('--T', type=float, default=80.0, help='Cell temperature [°C]')
    ap.add_argument('--p-cathode', type=float, default=0.0,
                    help='Cathode H₂ pressure [barg]')
    ap.add_argument('--p-anode', type=float, default=0.0,
                    help='Anode O₂ pressure [barg]')
    ap.add_argument('--save-dir', type=str, default=None,
                    help='Directory to save plots and data')
    ap.add_argument('--title', type=str, default=None, help='Plot title')
    args = ap.parse_args()

    eis_ref_v = None

    # ── Resolve inputs ──
    if args.folder:
        # Folder mode
        geo_area = args.area
        if geo_area is None:
            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0
        fp, eis_fps = scan_folder(args.folder, cell_id=args.cell_id)
        if fp is None:
            print("  ERROR: No polcurve file found.")
            return
        eis_ref_v = args.eis_ref_voltage
        if eis_ref_v is None and eis_fps:
            v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
            eis_ref_v = float(v_str) if v_str else 1.25
        save = args.save_dir
        T_C, p_cath, p_an = args.T, args.p_cathode, args.p_anode
        title = args.title

    elif args.file:
        # Direct file mode
        fp = _clean_path(args.file)
        eis_fps = [_clean_path(e) for e in args.eis] if args.eis else []
        geo_area = args.area
        if geo_area is None:
            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0
        eis_ref_v = args.eis_ref_voltage
        if eis_ref_v is None and eis_fps:
            v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
            eis_ref_v = float(v_str) if v_str else 1.25
        save = args.save_dir
        T_C, p_cath, p_an = args.T, args.p_cathode, args.p_anode
        title = args.title

    else:
        # Interactive mode
        print("=" * 60)
        print("  PEM Electrolyzer Polarization Curve Analyzer")
        print("=" * 60)
        print()

        mode = input("  [1] Single polcurve file  [2] Folder with polcurve + EIS  (2): ").strip() or '2'

        if mode == '2':
            folder = input("  Folder path: ").strip()
            folder = _clean_path(folder)
            cell_id = input("  Cell ID (a1): ").strip() or 'a1'
            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0

            fp, eis_fps = scan_folder(folder, cell_id=cell_id)
            if fp is None:
                print("  ERROR: No polcurve file found.")
                return

            if eis_fps:
                v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
                eis_ref_v = float(v_str) if v_str else 1.25
        else:
            fp = input("  Polcurve data file path: ").strip()
            fp = _clean_path(fp)
            if not fp:
                print("  No file provided.")
                return

            eis_input = input("  EIS data file path (Enter = none): ").strip()
            eis_fps = [_clean_path(eis_input)] if eis_input else []

            area_str = input("  Geometric electrode area [cm²] (5.0): ").strip()
            geo_area = float(area_str) if area_str else 5.0

            if eis_fps:
                v_str = input("  EIS reference voltage [V] for filtering (1.25): ").strip()
                eis_ref_v = float(v_str) if v_str else 1.25

        print("\n  Operating conditions for model fit (Enter = default):")
        t_str = input("    Temperature [°C] (80.0): ").strip()
        T_C = float(t_str) if t_str else 80.0
        pc_str = input("    Cathode pressure [barg] (0.0): ").strip()
        p_cath = float(pc_str) if pc_str else 0.0
        pa_str = input("    Anode pressure [barg] (0.0): ").strip()
        p_an = float(pa_str) if pa_str else 0.0

        save = input("\n  Save directory (Enter = display only): ").strip()
        save = _clean_path(save) if save else None
        title = input("  Plot title (Enter = auto): ").strip() or None
        print()

    analyze(fp, geo_area=geo_area, save_dir=save, title=title,
            T_C=T_C, p_cathode_barg=p_cath, p_anode_barg=p_an,
            eis_files=eis_fps if eis_fps else None,
            eis_ref_voltage=eis_ref_v)


if __name__ == '__main__':
    main()
