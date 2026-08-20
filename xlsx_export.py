"""
xlsx exports for the View tab.
===============================

Three scopes, all built from a stored detail record — no source files needed:

    build_sample_workbook(detail)             every analysis in the sample
    build_analysis_workbook(detail, bucket)   one characterization, all steps
    build_plot_workbook(detail, plot)         one plot, one sheet per panel

Layout follows the existing batch-Excel house style: a Summary sheet of scalars
per analysis, and a Data sheet of horizontal blocks — merged label row above
sub-headers — so an exported workbook is interchangeable with one the analysis
tab produced.

Data sheets are reconstructed from stored sidecars, which hold every plotted
series at the plotted resolution. Where a sidecar was not stored — cleaning, by
default, since its plots were 81 % of a run's payload — the Summary survives and
the Data sheet is replaced by a note saying so, rather than being silently absent.
"""

import io
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.helpers.record import decode_sidecars, json_safe, plot_bucket

FONT = 'Arial'
_HDR = Font(name=FONT, bold=True, size=10)
_LBL = Font(name=FONT, bold=True, size=11)
_BODY = Font(name=FONT, size=10)
_NOTE = Font(name=FONT, size=9, italic=True, color='808080')
_HFILL = PatternFill('solid', fgColor='D9E1F2')
_LFILL = PatternFill('solid', fgColor='E2EFDA')
_WFILL = PatternFill('solid', fgColor='FCE4D6')

# Buckets whose sidecars are excluded at write time keep their metrics but have
# no plot data to export. Stated in the file rather than left to be discovered.
NO_PLOT_DATA_NOTE = (
    'Plot data is not stored for this analysis — its sidecars are excluded at '
    'write time to stay under the transport limit. The summary above is '
    'complete; the per-point traces cannot be exported.')

_BAD_SHEET_CHARS = ':\\/?*[]'


def safe_sheet_name(name: str, used: Optional[set] = None) -> str:
    """Excel forbids : \\ / ? * [ ] in sheet names and caps them at 31 chars."""
    for ch in _BAD_SHEET_CHARS:
        name = name.replace(ch, '-')
    name = (name.strip() or 'Sheet')[:31]
    if used is None:
        return name
    base, n = name, 2
    while name in used:
        suffix = f' ({n})'
        name = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def _put(ws, r, c, v, font=_BODY):
    cell = ws.cell(row=r, column=c, value=json_safe(v))
    cell.font = font
    return cell


def _info_sheet(wb, title, rows, note=None):
    ws = wb.create_sheet(safe_sheet_name(title))
    _put(ws, 1, 1, title, _LBL)
    r = 3
    for k, v in rows:
        _put(ws, r, 1, k, _HDR)
        _put(ws, r, 2, v)
        r += 1
    if note:
        _put(ws, r + 1, 1, note, _NOTE).fill = _WFILL
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 62
    return ws


def _table_sheet(wb, title, headers, rows, used=None, note=None):
    ws = wb.create_sheet(safe_sheet_name(title, used))
    for i, h in enumerate(headers, start=1):
        cell = _put(ws, 1, i, h, _HDR)
        cell.fill = _HFILL
        cell.alignment = Alignment(horizontal='center')
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            _put(ws, r, c, v)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = \
            max(12, min(30, len(str(h)) + 4))
    if note:
        _put(ws, len(rows) + 3, 1, note, _NOTE).fill = _WFILL
    ws.freeze_panes = 'A2'
    return ws


def _block_sheet(wb, title, blocks, used=None, note=None):
    """Horizontal blocks side by side, each (label, [(header, values)])."""
    ws = wb.create_sheet(safe_sheet_name(title, used))
    col = 1
    longest = 0
    for label, cols in blocks:
        n = len(cols)
        cell = _put(ws, 1, col, label, _HDR)
        cell.fill = _LFILL
        if n > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + n - 1)
        for i, (head, vals) in enumerate(cols):
            c = _put(ws, 2, col + i, head, _HDR)
            c.fill = _HFILL
            c.alignment = Alignment(horizontal='center')
            for j, v in enumerate(vals):
                _put(ws, 3 + j, col + i, v)
            ws.column_dimensions[get_column_letter(col + i)].width = 17
            longest = max(longest, len(vals))
        col += n + 1
    if note:
        _put(ws, longest + 5, 1, note, _NOTE).fill = _WFILL
    ws.freeze_panes = 'A3'
    return ws


# ─────────────────────────────────────────────────────────────────────
#  Reading the stored record
# ─────────────────────────────────────────────────────────────────────

def _fmt_conditions(cond: Dict[str, Any]) -> str:
    if not cond:
        return ''
    bits = []
    if cond.get('T_C') is not None:
        bits.append(f"{cond['T_C']:g} °C")
    if cond.get('RH_pct') is not None:
        bits.append(f"{cond['RH_pct']:g} % RH")
    if cond.get('P_value') is not None:
        bits.append(f"{cond['P_value']:g} {cond.get('P_unit', '')}".strip())
    for k, lab in (('H2_slpm', 'H₂'), ('Air_slpm', 'Air'),
                   ('N2_slpm', 'N₂'), ('O2_slpm', 'O₂')):
        if cond.get(k) is not None:
            bits.append(f"{lab} {cond[k]:g} slpm")
    return ' · '.join(bits)


def _plots_of(detail: Dict[str, Any], bucket: Optional[str] = None
              ) -> List[Tuple[str, str, Dict[str, Any]]]:
    """[(bucket, plot_name, entry)] sorted by bucket then step then name."""
    out = []
    for b, plots in (detail.get('metrics') or {}).items():
        if bucket and b != bucket:
            continue
        for name, entry in plots.items():
            out.append((b, name, entry))
    # Aggregate views — batch overlays — carry no step, so a naive sort puts
    # them ahead of the measurements they summarise. Push them to the end.
    def key(t):
        step = str((t[2].get('conditions') or {}).get('step') or '')
        return (t[0], 0 if step else 1, step, t[1])
    out.sort(key=key)
    return out


def _summary_rows_for(detail: Dict[str, Any], bucket: str) -> List[Dict[str, Any]]:
    """Summary rows attributable to one analysis.

    Rows carry an Analysis when the Full Analysis orchestrator wrote them and
    only a Label otherwise; fall back to matching the label against the
    bucket's plot names, which is how those names are constructed.
    """
    rows = detail.get('summary') or []
    names = [n for _b, n, _e in _plots_of(detail, bucket)]
    out = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        if row.get('Analysis'):
            if row['Analysis'] == bucket:
                out.append(row)
            continue
        label = str(row.get('Label') or row.get('label') or '')
        if label and any(label in n for n in names):
            out.append(row)
    return out


def _summary_table(detail: Dict[str, Any], bucket: str):
    """(headers, rows) for a bucket's Summary sheet, union of all row keys."""
    rows = _summary_rows_for(detail, bucket)
    step_by_label = {}
    for _b, name, entry in _plots_of(detail, bucket):
        step = str((entry.get('conditions') or {}).get('step') or '')
        cond = _fmt_conditions(entry.get('conditions') or {})
        for row in rows:
            label = str(row.get('Label') or row.get('label') or '')
            if label and label in name:
                step_by_label.setdefault(label, (step, cond))

    keys: List[str] = []
    for row in rows:
        for k in row:
            if k in ('Label', 'label', 'Analysis') or k in keys:
                continue
            keys.append(k)

    headers = ['Label', 'Step', 'Conditions'] + keys
    table = []
    for row in rows:
        label = str(row.get('Label') or row.get('label') or '')
        step, cond = step_by_label.get(label, ('', ''))
        table.append([label, step, cond] + [row.get(k) for k in keys])
    return headers, table


def _panels(sidecar: Dict[str, Any]):
    """[(panel_title, xlabel, [(line_label, xs, ys)])] for panels with data."""
    out = []
    for ax in (sidecar.get('data') or {}).get('axes', []):
        lines = [(ln.get('label') or 'series', ln.get('x') or [], ln.get('y') or [])
                 for ln in ax.get('lines', []) if ln.get('x')]
        if lines:
            out.append((ax.get('title') or ax.get('ylabel') or 'Panel',
                        ax.get('xlabel') or 'x', lines))
    return out


def _readouts(sidecar: Dict[str, Any]) -> List[Tuple[str, str]]:
    vals = []
    for ax in (sidecar.get('data') or {}).get('axes', []):
        for t in ax.get('texts', []):
            for line in str(t.get('text') or '').split('\n'):
                if '=' in line:
                    k, v = line.split('=', 1)
                    vals.append((k.strip(), v.strip()))
        for rl in ax.get('axhlines', []) + ax.get('axvlines', []):
            lab = str(rl.get('label') or '')
            if '=' in lab:
                k, v = lab.split('=', 1)
                vals.append((k.strip(), v.strip()))
    return vals


def _data_blocks(detail, sidecars, bucket):
    """One block per plot in a bucket: its first panel's series."""
    blocks = []
    for _b, name, entry in _plots_of(detail, bucket):
        sc = sidecars.get(name)
        if not sc:
            continue
        panels = _panels(sc)
        if not panels:
            continue
        title, xlabel, lines = panels[0]
        cols = [(xlabel, lines[0][1])]
        for label, _x, ys in lines:
            cols.append((label, ys))
        # Later panels sharing this x get appended as extra columns.
        for ptitle, pxlabel, plines in panels[1:]:
            if pxlabel == xlabel and len(plines[0][2]) == len(lines[0][1]):
                for label, _x, ys in plines:
                    cols.append((f'{ptitle} · {label}' if label != 'series'
                                 else ptitle, ys))
        step = str((entry.get('conditions') or {}).get('step') or '')
        blocks.append((f'{name}' if not step else f'{step} · {name}', cols))
    return blocks


def _header_rows(detail, scope):
    return [
        ('Sample', detail.get('sample_name', '')),
        ('Script', detail.get('script', '')),
        ('Analysed', detail.get('timestamp', '')),
        ('Job', detail.get('job_id', '')),
        ('Contributing jobs',
         ', '.join(j.get('job_id', '') for j in (detail.get('jobs') or []))
         or detail.get('job_id', '')),
        ('Analyses present', ', '.join(sorted((detail.get('metrics') or {}).keys()))),
        ('Export scope', scope),
    ]


# ─────────────────────────────────────────────────────────────────────
#  Public builders
# ─────────────────────────────────────────────────────────────────────

def _finish(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_sample_workbook(detail: Dict[str, Any]) -> bytes:
    """Every analysis in the sample: Summary + Data per characterization."""
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    _info_sheet(wb, 'Sample Info', _header_rows(detail, 'Whole sample'),
                note='Reconstructed from stored analysis data.')

    sidecars = decode_sidecars(detail)
    for bucket in sorted((detail.get('metrics') or {}).keys()):
        headers, rows = _summary_table(detail, bucket)
        if rows:
            _table_sheet(wb, f'{bucket} Summary', headers, rows, used)
        blocks = _data_blocks(detail, sidecars, bucket)
        if blocks:
            _block_sheet(wb, f'{bucket} Data', blocks, used)
        else:
            ws = wb.create_sheet(safe_sheet_name(f'{bucket} Data', used))
            _put(ws, 1, 1, f'{bucket} — no plot data', _LBL)
            _put(ws, 3, 1, NO_PLOT_DATA_NOTE, _NOTE).fill = _WFILL
            ws.column_dimensions['A'].width = 110
    if not wb.sheetnames:
        wb.create_sheet('No Data')
    return _finish(wb)


def build_analysis_workbook(detail: Dict[str, Any], bucket: str) -> bytes:
    """One characterization, every step."""
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    _info_sheet(wb, 'Sample Info',
                _header_rows(detail, f'{bucket} only — all steps'))
    headers, rows = _summary_table(detail, bucket)
    if rows:
        _table_sheet(wb, f'{bucket} Summary', headers, rows, used)
    blocks = _data_blocks(detail, decode_sidecars(detail), bucket)
    if blocks:
        _block_sheet(wb, f'{bucket} Data', blocks, used)
    else:
        ws = wb.create_sheet(safe_sheet_name(f'{bucket} Data', used))
        _put(ws, 1, 1, f'{bucket} — no plot data', _LBL)
        _put(ws, 3, 1, NO_PLOT_DATA_NOTE, _NOTE).fill = _WFILL
        ws.column_dimensions['A'].width = 110
    return _finish(wb)


def build_plot_workbook(detail: Dict[str, Any], plot: str) -> bytes:
    """One plot: Plot Info, Readouts, then one sheet per panel.

    Panels get separate sheets because they do not share an x axis — current
    density against cycle index, say — so a combined sheet would need padding.
    """
    wb = Workbook()
    wb.remove(wb.active)
    used = set()

    entry, bucket = None, ''
    for b, name, e in _plots_of(detail):
        if name == plot:
            entry, bucket = e, b
            break
    cond = (entry or {}).get('conditions') or {}
    sc = decode_sidecars(detail).get(plot)
    panels = _panels(sc) if sc else []

    _info_sheet(wb, 'Plot Info', [
        ('Plot', plot),
        ('Analysis', bucket),
        ('Step', str(cond.get('step') or '')),
        ('Conditions', _fmt_conditions(cond)),
        ('Sample', detail.get('sample_name', '')),
        ('Analysed', detail.get('timestamp', '')),
        ('Panels', ', '.join(p[0] for p in panels) or '—'),
    ], note=None if panels else NO_PLOT_DATA_NOTE)

    values = (entry or {}).get('values') or {}
    ro = _readouts(sc) if sc else [
        (k, str(v.get('value')) + ' ' + str(v.get('unit', ''))
         if isinstance(v, dict) else str(v)) for k, v in values.items()]
    if ro:
        _table_sheet(wb, 'Readouts', ['Readout', 'Value'],
                     [[k, v] for k, v in ro], used)

    for title, xlabel, lines in panels:
        cols = [(xlabel, lines[0][1])]
        for label, _x, ys in lines:
            cols.append((label, ys))
        _block_sheet(wb, title, [(title, cols)], used)

    if not wb.sheetnames:
        wb.create_sheet('No Data')
    return _finish(wb)


def export_filename(detail: Dict[str, Any], bucket: str = '',
                    plot: str = '') -> str:
    sample = str(detail.get('sample_name') or 'export')
    keep = ''.join(c if c.isalnum() or c in '-_.' else '_' for c in sample)
    if plot:
        pk = ''.join(c if c.isalnum() or c in '-_.' else '_' for c in plot)
        return f'{keep}_{pk}.xlsx'[:120]
    if bucket:
        return f'{keep}_{bucket}.xlsx'[:120]
    return f'{keep}.xlsx'[:120]
